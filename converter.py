import os
from openpyxl import load_workbook

# '$' signals a lower octave
# '&' signals a natural

CONVERSION_TABLE = {
    "$F#": 0,
    "$G": 1,
    "$G#": 2,
    "$A": 3,
    "$A#": 4,
    "$B": 5,
    "$C": 6,
    "$C#": 7,
    "$D": 8,
    "$D#": 9,
    "$E": 10,
    "$F": 11,
    "$F#": 12,
    "G": 13,
    "G#": 14,
    "A": 15,
    "A#": 16,
    "B": 17,
    "C": 18,
    "C#": 19,
    "D": 20,
    "D#": 21,
    "E": 22,
    "F": 23,
    "F#": 24,
    "-": 99
}

# CHANGE THESE ACCORDING TO YOUR OWN WORLD
COORDINATES = {
    0: [-1, -59, -1],
    1: [1, -59, -1],
    2: [3, -59, -1],
    3: [5, -59, -1],
    4: [7, -59, -1],
    5: [9, -59, -1],
    6: [11, -59, -1],
    7: [13, -59, -1],
    8: [15, -59, -1],
    9: [17, -59, -1],
    10: [21, -59, -1],
    11: [23, -59, -1],
    12: [25, -59, -1],
    13: [27, -59, -1],
    14: [29, -59, -1],
    15: [31, -59, -1],
    16: [33, -59, -1],
    17: [35, -59, -1],
    18: [37, -59, -1],
    19: [39, -59, -1],
    20: [41, -59, -1],
    21: [43, -59, -1],
    22: [45, -59, -1],
    23: [47, -59, -1],
    24: [49, -59, -1],
    99: [51, -59, -1],
}

class Converter:
    def __init__(self, filename, scale, scaletype):
        self.filename = filename
        self.scale = scale
        self.scaletype = scaletype

    def get_notes(self, sheet_name):
        workbook = load_workbook(self.filename, sheet_name)
        sheet = workbook[sheet_name]

        notes = []
        for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
            value = row[0]  # Get the value in column A
            if value is not None:
                notes.append(value)

        return notes
    
    def convert_notes_to_numbers(self, notes):
        toReturn = []
        for note in notes:
            corresponding_number = CONVERSION_TABLE[note.replace("&", "")]
            if note.replace("$", "") in self.scale:
                if self.scaletype == "flat":
                    corresponding_number -= 1
                elif self.scaletype == "sharp":
                    corresponding_number += 1
            toReturn.append(corresponding_number)
        return toReturn
    
    def return_five_bit_binary(self, n):
        return self.do_five_bits(n, [], 4)

    def do_five_bits(self, n, bits, digit):
        if digit < 0:
            return bits
        else:
            if n - 2 ** digit >= 0:
                bits.append(1)
                return self.do_five_bits(n - 2 ** digit, bits, digit - 1)
            else:
                bits.append(0)
                return self.do_five_bits(n, bits, digit - 1)
    
    def place_block_command(self, coords, blockname):
        return f"/setblock {coords[0]} {coords[1]} {coords[2]} minecraft:{blockname}"
    
    def write_commands(self, note_numbers, foldername):
        if not os.path.exists(foldername):
            os.makedirs(foldername)

        # Iterate over note_numbers and write each function{i}.mcfunction file
        for i, num in enumerate(note_numbers):
            # Create the file path for the current function
            function_filename = os.path.join(foldername, f"function{i}.mcfunction")

            # Open the file for writing
            with open(function_filename, 'w') as f:
                # Build the commands for the current function
                if num <= 25:
                    # Place a redstone block
                    f.write(self.place_block_command(COORDINATES[num], "redstone_block") + "\n")
                else:
                    # Place air if the note number is out of range
                    f.write(self.place_block_command(COORDINATES[num], "air") + "\n")
                
                # Always place air after placing the block
                f.write(self.place_block_command(COORDINATES[num], "air") + "\n")
                
                # Schedule the next function to run after 1 tick, if it's not the last function
                if i + 1 < len(note_numbers):
                    next_function = f"function{i+1}"
                    f.write(f"schedule function my_namespace:{next_function} 1t\n")

        print(f"Commands written to folder '{foldername}'.")

        

    def print_commands(self, note_numbers):
        for num in note_numbers:
            if num <= 25:
                print(self.place_block_command(COORDINATES[num], "redstone_block"))
            else:
                print(self.place_block_command(COORDINATES[num], "air"))
            print(self.place_block_command(COORDINATES[num], "air"))

if __name__ == '__main__':
    c = Converter("songs.xlsx", ['B', 'E'], "flat") # b flat major
    notes = c.get_notes("SuperIdol")
    print(notes)
    numbers = c.convert_notes_to_numbers(notes)
    print(numbers)
    c.print_commands(numbers)

    c.write_commands(numbers, "testdir")